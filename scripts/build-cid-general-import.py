#!/usr/bin/env python3
"""Normalize the CID General workbook + extracted photo manifest for import.

Requires openpyxl. The generated payload is operational data and should remain
outside git. It is deliberately conservative: blank template rows are ignored,
and verification is never upgraded beyond what the workbook states.
"""
import argparse, hashlib, json, re
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

SPECIAL = {"Template", "Weapons Crafting Benches", "Drug Crafting Benches", "MIsc Criminal Areas"}
TYPE = {"Weapons Crafting Benches": "stash_house", "Drug Crafting Benches": "drug_lab", "MIsc Criminal Areas": "dead_drop"}

def clean(v):
    if v is None: return None
    if isinstance(v, (datetime, date)): return v.isoformat()
    if isinstance(v, float) and v.is_integer(): return str(int(v))
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s or None

def threat(v):
    v = (clean(v) or "low").lower()
    return v if v in {"low", "medium", "high", "critical"} else "low"

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path)
    ap.add_argument("photo_manifest", type=Path)
    ap.add_argument("--photos", type=Path, required=True)
    ap.add_argument("--output", type=Path, required=True)
    a = ap.parse_args()
    # Normal mode gives reliable dimensions for Google-Sheets-derived XLSX
    # files; read-only mode reports None and repeatedly re-streams sparse XML.
    wb = load_workbook(a.workbook, data_only=True, read_only=False)
    manifest = json.loads(a.photo_manifest.read_text(encoding="utf-8"))
    out = {"source": "CID General / Gang Fact Sheet", "gangs": [], "members": [], "turf": [], "places": [], "media": []}
    gang_sheet_to_key = {}
    for ws in wb.worksheets:
        if ws.title in SPECIAL: continue
        # Standard sheets start their headings on row 1; DRENGR has one junk row.
        header = 1 if clean(ws.cell(1, 2).value) == "Gang Name" else 2 if clean(ws.cell(2, 2).value) in {"Gang Name", "C"} else None
        if not header: continue
        name = clean(ws.cell(header + 1, 2).value)
        if not name or name == "Gang Name": continue
        key = ws.title.strip(); gang_sheet_to_key[ws.title] = key
        classification = clean(ws.cell(header + 1, 7).value)
        aliases = [] if clean(ws.title).lower() == name.lower() else [clean(ws.title)]
        # A few legacy titles encode aliases as "Name or Alias". Feed both
        # sides to the matcher so they resolve to an existing canonical row.
        if re.search(r"\s+or\s+", name, flags=re.I):
            aliases.extend(clean(x) for x in re.split(r"\s+or\s+", name, flags=re.I))
        aliases = list(dict.fromkeys(x for x in aliases if x and x.lower() != name.lower()))
        out["gangs"].append({"source_key": key, "name": name, "aliases_list": aliases, "aliases": ", ".join(aliases) or None, "colors": clean(ws.cell(header + 2, 4).value), "classification": classification.lower().replace(" ", "_") if classification else None, "threat_level": threat(ws.cell(header + 1, 9).value), "status": "active", "confidence": "unverified", "notes": f"Workbook last updated: {clean(ws.cell(header + 4, 2).value) or 'unknown'}; workbook lead: {clean(ws.cell(header + 4, 9).value) or 'unassigned'}"})
        territory = clean(ws.cell(header + 1, 21).value)
        if territory and territory.lower() not in {"unsure", "tbd"}: out["turf"].append({"gang_key": key, "block": territory, "confidence": "unverified"})
        member_header = header + 6
        # Some Google-Sheets-derived tabs omit worksheet dimensions in the
        # XLSX XML. Member tables are bounded, so use a conservative fallback.
        for r in range(member_header + 1, (ws.max_row or 200) + 1):
            member = clean(ws.cell(r, 2).value)
            if not member: continue
            rank = clean(ws.cell(r, 3).value)
            if member.lower() in {"name", "gang name", "last updated"}: continue
            out["members"].append({"gang_key": key, "name": member, "rank": rank, "note": clean(ws.cell(r, 4).value), "mugshot_url": clean(ws.cell(r, 12).value) if isinstance(ws.cell(r, 12).value, str) and str(ws.cell(r, 12).value).startswith("http") else None, "confidence": "unverified", "ccw": bool(ws.cell(r, 15).value) if ws.cell(r, 15).value is not None else None, "vch": 1 if ws.cell(r, 14).value is True else None})
    place_keys = {}
    for sheet in TYPE:
        ws = wb[sheet]
        for r in range(3, (ws.max_row or 500) + 1):
            name, area, notes = clean(ws.cell(r, 2).value), clean(ws.cell(r, 3).value), clean(ws.cell(r, 7).value)
            if not name or name.lower() in {"misc criminal areas", "weapons crafting benches"}: continue
            key = f"{sheet}:{r}"; place_keys[(sheet, r)] = key
            out["places"].append({"source_key": key, "name": name, "area": area, "type": TYPE[sheet], "notes": notes or f"Imported from {sheet}; verification not specified."})
    for item in manifest:
        src = a.photos / item["file"]
        if not src.exists(): raise FileNotFoundError(src)
        sheet, row = item["sheet"], int(item["row"])
        gang_key = gang_sheet_to_key.get(sheet)
        place_key = place_keys.get((sheet, row))
        nearby = item.get("nearby") or []
        out["media"].append({"file": str(src.resolve()), "title": f"{sheet} — {clean(nearby[0]) or item['file']}", "sheet": sheet, "anchor": f"R{row}C{item['column']}", "sha256": item.get("sha256") or hashlib.sha256(src.read_bytes()).hexdigest(), "gang_key": gang_key, "place_key": place_key, "category": "places" if place_key else "other", "labels": ["CID General import", "Gang attire" if gang_key else "Location intelligence"], "confidence": "unverified"})
    a.output.parent.mkdir(parents=True, exist_ok=True)
    a.output.write_text(json.dumps(out, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({k: len(v) for k, v in out.items() if isinstance(v, list)}, indent=2))

if __name__ == "__main__": main()
